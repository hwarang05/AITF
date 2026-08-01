"""
XLSX Parser
"""

from pathlib import Path

from openpyxl import load_workbook

from app.parsers.base import BaseParser


class XlsxParser(BaseParser):
    """
    XLSX Parser
    """

    extensions = {
        ".xlsx",
        ".xlsm",
    }

    def supports(
        self,
        path: Path,
    ) -> bool:

        return path.suffix.lower() in self.extensions

    def parse(
        self,
        path: Path,
    ) -> str:

        workbook = load_workbook(
            filename=path,
            data_only=True,
        )

        texts: list[str] = []

        for sheet in workbook.worksheets:

            texts.append(f"[Sheet] {sheet.title}")

            for row in sheet.iter_rows(values_only=True):

                values = [
                    str(value)
                    for value in row
                    if value is not None
                ]

                if values:
                    texts.append(" | ".join(values))

        workbook.close()

        return "\n".join(texts)